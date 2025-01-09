from datetime import datetime
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import utils

# Define form variables
school_var = tk.StringVar()
date_var = tk.StringVar()
name_var = tk.StringVar()
birthday_var = tk.StringVar()
phone_number_var = tk.StringVar()

distance_vision_var = tk.StringVar()
distance_vision_OD_var = tk.StringVar()
distance_vision_OI_var = tk.StringVar()

intraocular_pressure_var = tk.StringVar()
intraocular_pressure_OD_var = tk.StringVar()
intraocular_pressure_OI_var = tk.StringVar()
intraocular_pressure_timestamp_OD_var = tk.StringVar()
intraocular_pressure_timestamp_OI_var = tk.StringVar()

dilatation_var = tk.BooleanVar()
dilatation_timestamp_var = tk.StringVar()

refractive_status_method_var = tk.StringVar()
refractive_status_OD_var = tk.StringVar()
refractive_status_OD_dVA_var = tk.StringVar()
refractive_status_OI_var = tk.StringVar()
refractive_status_OI_dVA_var = tk.StringVar()

lens_prescription_OD_var = tk.StringVar()
lens_prescription_OI_var = tk.StringVar()

eyelid_OD_var = tk.StringVar()
eyelid_OI_var = tk.StringVar()
eyelid_OD_var.set("Normal")
eyelid_OI_var.set("Normal")
eyelid_OD_comments_var = tk.StringVar()
eyelid_OI_comments_var = tk.StringVar()

conjunctiva_OD_var = tk.StringVar()
conjunctiva_OI_var = tk.StringVar()
conjunctiva_OD_var.set("Normal")
conjunctiva_OI_var.set("Normal")
conjunctiva_OD_comments_var = tk.StringVar()
conjunctiva_OI_comments_var = tk.StringVar()

cornea_OD_var = tk.StringVar()
cornea_OI_var = tk.StringVar()
cornea_OD_var.set("Normal")
cornea_OI_var.set("Normal")
cornea_OD_comments_var = tk.StringVar()
cornea_OI_comments_var = tk.StringVar()

iris_OD_var = tk.StringVar()
iris_OI_var = tk.StringVar()
iris_OD_var.set("Normal")
iris_OI_var.set("Normal")
iris_OD_comments_var = tk.StringVar()
iris_OI_comments_var = tk.StringVar()

pupil_OD_var = tk.StringVar()
pupil_OI_var = tk.StringVar()
pupil_OD_var.set("Normal")
pupil_OI_var.set("Normal")
pupil_OD_comments_var = tk.StringVar()
pupil_OI_comments_var = tk.StringVar()

lente_OD_var = tk.StringVar()
lente_OI_var = tk.StringVar()
lente_OD_var.set("Normal")
lente_OI_var.set("Normal")
lente_OD_comments_var = tk.StringVar()
lente_OI_comments_var = tk.StringVar()

cd_OD_var = tk.StringVar()
cd_OI_var = tk.StringVar()

nerve_OD_var = tk.StringVar()
nerve_OI_var = tk.StringVar()
nerve_OD_var.set("Normal")
nerve_OI_var.set("Normal")
nerve_OD_comments_var = tk.StringVar()
nerve_OI_comments_var = tk.StringVar()

macula_OD_var = tk.StringVar()
macula_OI_var = tk.StringVar()
macula_OD_var.set("Normal")
macula_OI_var.set("Normal")
macula_OD_comments_var = tk.StringVar()
macula_OI_comments_var = tk.StringVar()

vasculature_OD_var = tk.StringVar()
vasculature_OI_var = tk.StringVar()
vasculature_OD_var.set("Normal")
vasculature_OI_var.set("Normal")
vasculature_OD_comments_var = tk.StringVar()
vasculature_OI_comments_var = tk.StringVar()

periphery_OD_var = tk.StringVar()
periphery_OI_var = tk.StringVar()
periphery_OD_var.set("Normal")
periphery_OI_var.set("Normal")
periphery_OD_comments_var = tk.StringVar()
periphery_OI_comments_var = tk.StringVar()

vitreous_OD_var = tk.StringVar()
vitreous_OI_var = tk.StringVar()
vitreous_OD_var.set("Normal")
vitreous_OI_var.set("Normal")
vitreous_OD_comments_var = tk.StringVar()
vitreous_OI_comments_var = tk.StringVar()

dx_myopia_var = tk.BooleanVar()
dx_hyperopia_var = tk.BooleanVar()
dx_astigmatism_var = tk.BooleanVar()

creation_date_var = tk.StringVar()
doctor_firm_var = tk.StringVar()
doctor_name_var = tk.StringVar()

headers = ["Clínica/Escuela", "Fecha", "Nombre", "Fecha_de_nacimiento", "Número_de_teléfono", "Razón_por_cita", 
           "Visión_distancia", "Visión_distancia_OD", "Visión_distancia_OI", 
            "Presión_intraocular", "Presión_intraocular_OD", "Presión_intraocular_OD_tiempo", "Presión_intraocular_OI", "Presión_intraocular_OI_tiempo", 
            "Dilatación", "Dilatación_tiempo",
            "Autorefraction/Retinoscopia", "Autorefraction/Retinoscopia_OD", "Autorefraction/Retinoscopia_OD_dVA", "Autorefraction/Retinoscopia_OI", "Autorefraction/Retinoscopia_OI_dVA",
            "Rx_de_anteojos_OD", "Rx_de_anteojos_OI",
            "Párpados_OD", "Párpados_OD_comentarios", "Párpados_OI", "Párpados_OI_comentarios",
            "Conjunctiva_OD", "Conjunctiva_OD_comentarios", "Conjunctiva_OI", "Conjunctiva_OI_comentarios",
            "Cornea_OD", "Cornea_OD_comentarios", "Cornea_OI", "Cornea_OI_comentarios",
            "Iris_OD", "Iris_OD_comentarios", "Iris_OI", "Iris_OI_comentarios",
            "Pupil_OD", "Pupil_OD_comentarios", "Pupil_OI", "Pupil_OI_comentarios",
            "Lente_OD", "Lente_OD_comentarios", "Lente_OI", "Lente_OI_comentarios",
            "C/D_OD", "C/D_OI", 
            "Nervio_OD", "Nervio_OD_comentarios", "Nervio_OI", "Nervio_OI_comentarios",
            "Macula_OD", "Macula_OD_comentarios", "Macula_OI", "Macula_OI_comentarios",
            "Vasculatura_OD", "Vasculatura_OD_comentarios", "Vasculatura_OI", "Vasculatura_OI_comentarios",
            "Periferia_OD", "Periferia_OD_comentarios", "Periferia_OI", "Periferia_OI_comentarios",
            "Vitreous_OD", "Vitreous_OD_comentarios", "Vitreous_OI", "Vitreous_OI_comentarios",
            "Dx_Miopía", "Dx_Hipermetropía", "Dx_Astigmatismo", "Dx_comentarios", "Tx",
            "Firme_de_Doctor", "Nombre_de_Doctor", "Fecha_de_creación", "Fecha_de_la_última_edición", "Addendum"]

def open_file(file_path): 
    # Create or load the workbook
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Student Data"
        # Add headers to the sheet
        sheet.append(headers)
        workbook.save(file_path)
        return workbook
    else:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        return workbook

def clear_form(mmhg_label_od, mmhg_label_oi, 
               dilatation_checkbox, reason_for_visit_textbox, 
               dx_comentarios_textbox, tx_textbox, 
               eyelid_OD_comment, eyelid_OI_comment, 
               conjunctiva_OD_comment, conjunctiva_OI_comment, 
               cornea_OD_comment, cornea_OI_comment, 
               iris_OD_comment, iris_OI_comment, 
               pupil_OD_comment, pupil_OI_comment, 
               lente_OD_comment, lente_OI_comment, 
               nerve_OD_comment, nerve_OI_comment, 
               macula_OD_comment, macula_OI_comment, 
               vasculature_OD_comment, vasculature_OI_comment, 
               periphery_OD_comment, periphery_OI_comment, 
               vitreous_OD_comment, vitreous_OI_comment): 
    school_var.set("")
    date_var.set("")
    name_var.set("")
    birthday_var.set("")
    phone_number_var.set("")
    reason_for_visit_textbox.delete("1.0", tk.END)

    distance_vision_var.set("")
    distance_vision_OD_var.set("")
    distance_vision_OI_var.set("")

    intraocular_pressure_var.set("")
    intraocular_pressure_OD_var.set("")
    intraocular_pressure_OI_var.set("")
    intraocular_pressure_timestamp_OD_var.set("")
    intraocular_pressure_timestamp_OI_var.set("")
    utils.update_text_input_timestamp(text_var=intraocular_pressure_OD_var, label=mmhg_label_od, label_text="mmHg", timestamp_var=intraocular_pressure_timestamp_OD_var)
    utils.update_text_input_timestamp(text_var=intraocular_pressure_OI_var, label=mmhg_label_oi, label_text="mmHg", timestamp_var=intraocular_pressure_timestamp_OI_var)

    dilatation_var.set(False)
    dilatation_timestamp_var.set("")
    utils.update_checkbox_timestamp(dilatation_timestamp_var, dilatation_checkbox, dilatation_var, "Dilatación")

    refractive_status_method_var.set("")
    refractive_status_OD_var.set("")
    refractive_status_OD_dVA_var.set("")
    refractive_status_OI_var.set("")
    refractive_status_OI_dVA_var.set("")

    lens_prescription_OD_var.set("")
    lens_prescription_OI_var.set("")

    eyelid_OD_var.set("Normal")
    eyelid_OI_var.set("Normal")
    eyelid_OD_comments_var.set("")
    eyelid_OI_comments_var.set("")
    eyelid_OD_comment.config(state="disabled")
    eyelid_OI_comment.config(state="disabled")

    conjunctiva_OD_var.set("Normal")
    conjunctiva_OI_var.set("Normal")
    conjunctiva_OD_comments_var.set("")
    conjunctiva_OI_comments_var.set("")
    conjunctiva_OD_comment.config(state="disabled")
    conjunctiva_OI_comment.config(state="disabled")

    cornea_OD_var.set("Normal")
    cornea_OI_var.set("Normal")
    cornea_OD_comments_var.set("")
    cornea_OI_comments_var.set("")
    cornea_OD_comment.config(state="disabled")
    cornea_OI_comment.config(state="disabled")

    iris_OD_var.set("Normal")
    iris_OI_var.set("Normal")
    iris_OD_comments_var.set("")
    iris_OI_comments_var.set("")
    iris_OD_comment.config(state="disabled")
    iris_OI_comment.config(state="disabled")

    pupil_OD_var.set("Normal")
    pupil_OI_var.set("Normal")
    pupil_OD_comments_var.set("")
    pupil_OI_comments_var.set("")
    pupil_OD_comment.config(state="disabled")
    pupil_OI_comment.config(state="disabled")

    lente_OD_var.set("Normal")
    lente_OI_var.set("Normal")
    lente_OD_comments_var.set("")
    lente_OI_comments_var.set("")
    lente_OD_comment.config(state="disabled")
    lente_OI_comment.config(state="disabled")

    cd_OD_var.set("")
    cd_OI_var.set("")

    nerve_OD_var.set("Normal")
    nerve_OI_var.set("Normal")
    nerve_OD_comments_var.set("")
    nerve_OI_comments_var.set("")
    nerve_OD_comment.config(state="disabled")
    nerve_OI_comment.config(state="disabled")

    macula_OD_var.set("Normal")
    macula_OI_var.set("Normal")
    macula_OD_comments_var.set("")
    macula_OI_comments_var.set("")
    macula_OD_comment.config(state="disabled")
    macula_OI_comment.config(state="disabled")

    vasculature_OD_var.set("Normal")
    vasculature_OI_var.set("Normal")
    vasculature_OD_comments_var.set("")
    vasculature_OI_comments_var.set("")
    vasculature_OD_comment.config(state="disabled")
    vasculature_OI_comment.config(state="disabled")

    periphery_OD_var.set("Normal")
    periphery_OI_var.set("Normal")
    periphery_OD_comments_var.set("")
    periphery_OI_comments_var.set("")
    periphery_OD_comment.config(state="disabled")
    periphery_OI_comment.config(state="disabled")

    vitreous_OD_var.set("Normal")
    vitreous_OI_var.set("Normal")
    vitreous_OD_comments_var.set("")
    vitreous_OI_comments_var.set("")
    vitreous_OD_comment.config(state="disabled")
    vitreous_OI_comment.config(state="disabled")

    dx_myopia_var.set(False)
    dx_hyperopia_var.set(False)
    dx_astigmatism_var.set(False)
    dx_comentarios_textbox.delete("1.0", tk.END)
    tx_textbox.delete("1.0", tk.END)

    creation_date_var.set("")
    doctor_firm_var.set("")
    doctor_name_var.set("")

def save_to_excel(workbook, file_path, mmhg_label_od, mmhg_label_oi, 
                  dilatation_checkbox, reason_for_visit_textbox, 
                  dx_comentarios_textbox, tx_textbox, 
                  eyelid_OD_comment, eyelid_OI_comment, 
                  conjunctiva_OD_comment, conjunctiva_OI_comment,
                  cornea_OD_comment, cornea_OI_comment, 
                  iris_OD_comment, iris_OI_comment, 
                  pupil_OD_comment, pupil_OI_comment, 
                  lente_OD_comment, lente_OI_comment, 
                  nerve_OD_comment, nerve_OI_comment, 
                  macula_OD_comment, macula_OI_comment, 
                  vasculature_OD_comment, vasculature_OI_comment, 
                  periphery_OD_comment, periphery_OI_comment, 
                  vitreous_OD_comment, vitreous_OI_comment, addendum_textbox = None, addendum_doctor_name_entry = None, addendum_date = "", previous_addendum = "",
                  row_index=-1, window=None):
    """Saves form data to the Excel file."""
    school = school_var.get()
    date = date_var.get()
    name = name_var.get()
    birthday = birthday_var.get()
    phone_number = phone_number_var.get()
    reason_for_visit = reason_for_visit_textbox.get("1.0", tk.END).strip()

    distance_vision = distance_vision_var.get()
    distance_vision_OD = distance_vision_OD_var.get()
    distance_vision_OI = distance_vision_OI_var.get()

    intraocular_pressure = intraocular_pressure_var.get()
    intraocular_pressure_OD = intraocular_pressure_OD_var.get()
    intraocular_pressure_OI = intraocular_pressure_OI_var.get()
    intraocular_pressure_OD_timestamp = intraocular_pressure_timestamp_OD_var.get()
    intraocular_pressure_OI_timestamp = intraocular_pressure_timestamp_OI_var.get()

    dilatation = dilatation_var.get()
    dilatation_timestamp = dilatation_timestamp_var.get()

    refractive_status_method = refractive_status_method_var.get()
    refractive_status_OD = refractive_status_OD_var.get()
    refractive_status_OD_dVA = refractive_status_OD_dVA_var.get()
    refractive_status_OI = refractive_status_OI_var.get()
    refractive_status_OI_dVA = refractive_status_OI_dVA_var.get()

    lens_prescription_OD = lens_prescription_OD_var.get()
    lens_prescription_OI = lens_prescription_OI_var.get()

    eyelid_OD = eyelid_OD_var.get()
    eyelid_OI = eyelid_OI_var.get()
    eyelid_OD_comentarios = eyelid_OD_comments_var.get()
    eyelid_OI_comentarios = eyelid_OI_comments_var.get()

    conjunctiva_OD = conjunctiva_OD_var.get()
    conjunctiva_OI = conjunctiva_OI_var.get()
    conjunctiva_OD_comentarios = conjunctiva_OD_comments_var.get()
    conjunctiva_OI_comentarios = conjunctiva_OI_comments_var.get()

    cornea_OD = cornea_OD_var.get()
    cornea_OI = cornea_OI_var.get()
    cornea_OD_comentarios = cornea_OD_comments_var.get()
    cornea_OI_comentarios = cornea_OI_comments_var.get()

    iris_OD = iris_OD_var.get()
    iris_OI = iris_OI_var.get()
    iris_OD_comentarios = iris_OD_comments_var.get()
    iris_OI_comentarios = iris_OI_comments_var.get()

    pupil_OD = pupil_OD_var.get()
    pupil_OI = pupil_OI_var.get()
    pupil_OD_comentarios = pupil_OD_comments_var.get()
    pupil_OI_comentarios = pupil_OI_comments_var.get()

    lente_OD = lente_OD_var.get()
    lente_OI = lente_OI_var.get()
    lente_OD_comentarios = lente_OD_comments_var.get()
    lente_OI_comentarios = lente_OI_comments_var.get()

    cd_OD = cd_OD_var.get()
    cd_OI = cd_OI_var.get()

    nerve_OD = nerve_OD_var.get()
    nerve_OI = nerve_OI_var.get()
    nerve_OD_comentarios = nerve_OD_comments_var.get()
    nerve_OI_comentarios = nerve_OI_comments_var.get()

    macula_OD = macula_OD_var.get()
    macula_OI = macula_OI_var.get()
    macula_OD_comentarios = macula_OD_comments_var.get()
    macula_OI_comentarios = macula_OI_comments_var.get()

    vasculature_OD = vasculature_OD_var.get()
    vasculature_OI = vasculature_OI_var.get()
    vasculature_OD_comentarios = vasculature_OD_comments_var.get()
    vasculature_OI_comentarios = vasculature_OI_comments_var.get()

    periphery_OD = periphery_OD_var.get()
    periphery_OI = periphery_OI_var.get()
    periphery_OD_comentarios = periphery_OD_comments_var.get()
    periphery_OI_comentarios = periphery_OI_comments_var.get()

    vitreous_OD = vitreous_OD_var.get()
    vitreous_OI = vitreous_OI_var.get()
    vitreous_OD_comentarios = vitreous_OD_comments_var.get()
    vitreous_OI_comentarios = vitreous_OI_comments_var.get()

    dx_myopia = dx_myopia_var.get()
    dx_hyperopia = dx_hyperopia_var.get()
    dx_astigmatism = dx_astigmatism_var.get()
    dx_comentarios = dx_comentarios_textbox.get("1.0", tk.END).strip()
    tx = tx_textbox.get("1.0", tk.END).strip()
    
    doctor_firm = doctor_firm_var.get()
    doctor_name = doctor_name_var.get()
    creation_date = creation_date_var.get()
    last_edit_date = datetime.now().strftime("%Y:%m:%d %H:%M:%S")

    sheet = workbook.active
    if row_index == -1: 
        if not school or not date or not name or not birthday or not phone_number:
            messagebox.showerror("Error", "Please fill in all required student information fields.")
            return

        if not distance_vision or not distance_vision_OD or not distance_vision_OI: 
            messagebox.showerror("Error", "Please fill in all required distance vision fields.")
            return
        
        if not intraocular_pressure or not intraocular_pressure_OD or not intraocular_pressure_OI: 
            messagebox.showerror("Error", "Please fill in all required intraocular pressure fields.")
            return
        
        if not refractive_status_method or not refractive_status_OD or not refractive_status_OD_dVA or not refractive_status_OI or not refractive_status_OI_dVA: 
            messagebox.showerror("Error", "Please fill in all required refractive status fields.")
            return
        
        if not eyelid_OD or not eyelid_OI or not conjunctiva_OD or not conjunctiva_OI or not cornea_OD or not cornea_OI or not iris_OD or not iris_OI or not pupil_OD or not pupil_OI or not lente_OD or not lente_OI or not nerve_OD or not nerve_OI or not macula_OD or not macula_OI or not vasculature_OD or not vasculature_OI or not periphery_OD or not periphery_OI or not vitreous_OD or not vitreous_OI: 
            messagebox.showerror("Error", "Please fill in all required fields.")
            return

        if not doctor_name or not doctor_firm: 
            messagebox.showerror("Error", "Please fill in all required doctor information fields.")
            return

        sheet.append([school, date, name, birthday, phone_number, reason_for_visit, 
                    distance_vision, distance_vision_OD, distance_vision_OI, 
                    intraocular_pressure, intraocular_pressure_OD, intraocular_pressure_OD_timestamp, intraocular_pressure_OI, intraocular_pressure_OI_timestamp, 
                    dilatation, dilatation_timestamp, 
                    refractive_status_method, refractive_status_OD, refractive_status_OD_dVA, refractive_status_OI, refractive_status_OI_dVA, 
                    lens_prescription_OI, lens_prescription_OD, 
                    eyelid_OD, eyelid_OD_comentarios, eyelid_OI, eyelid_OI_comentarios, 
                    conjunctiva_OD, conjunctiva_OD_comentarios, conjunctiva_OI, conjunctiva_OI_comentarios,
                    cornea_OD, cornea_OD_comentarios, cornea_OI, cornea_OI_comentarios, 
                    iris_OD, iris_OD_comentarios, iris_OI, iris_OI_comentarios, 
                    pupil_OD, pupil_OD_comentarios, pupil_OI, pupil_OI_comentarios, 
                    lente_OD, lente_OD_comentarios, lente_OI, lente_OI_comentarios, 
                    cd_OD, cd_OI, 
                    nerve_OD, nerve_OD_comentarios, nerve_OI, nerve_OI_comentarios, 
                    macula_OD, macula_OD_comentarios, macula_OI, macula_OI_comentarios, 
                    vasculature_OD, vasculature_OD_comentarios, vasculature_OI, vasculature_OI_comentarios, 
                    periphery_OD, periphery_OD_comentarios, periphery_OI, periphery_OI_comentarios, 
                    vitreous_OD, vitreous_OD_comentarios, vitreous_OI, vitreous_OI_comentarios, 
                    dx_myopia, dx_hyperopia, dx_astigmatism, dx_comentarios, tx, 
                    doctor_firm, doctor_name, creation_date, last_edit_date, ""])
        workbook.save(file_path)
        messagebox.showinfo("Success", "Data saved successfully!")
        clear_form(mmhg_label_od, mmhg_label_oi, dilatation_checkbox, 
               reason_for_visit_textbox, dx_comentarios_textbox, tx_textbox, 
               eyelid_OD_comment, eyelid_OI_comment, 
               conjunctiva_OD_comment, conjunctiva_OI_comment, 
               cornea_OD_comment, cornea_OI_comment, 
               iris_OD_comment, iris_OI_comment, 
               pupil_OD_comment, pupil_OI_comment, 
               lente_OD_comment, lente_OI_comment, 
               nerve_OD_comment, nerve_OI_comment, 
               macula_OD_comment, macula_OI_comment, 
               vasculature_OD_comment, vasculature_OI_comment, 
               periphery_OD_comment, periphery_OI_comment, 
               vitreous_OD_comment, vitreous_OI_comment)
    else: 
        addendum = ""
        addendum_text = addendum_textbox.get("1.0", tk.END).strip()
        addendum_doctor_name = addendum_doctor_name_entry.get().strip()
        if addendum_text != "":
            if addendum_doctor_name != "":
                addendum = f"On {addendum_date}, Dr. {addendum_doctor_name} added: {addendum_text}"
            else: 
                messagebox.showerror("Error", "Please fill out all fields.")
                return
        
            new_addendum = previous_addendum + "\n" + addendum

            excel_row = row_index + 1 
            sheet.cell(row=excel_row, column=78, value=last_edit_date)
            sheet.cell(row=excel_row, column=79, value=new_addendum)
            workbook.save(file_path)
            messagebox.showinfo("Success", "Data updated successfully!")
        window.destroy()    

def create_form(workbook, content_frame, file_path, row):
    ######################## student information ###############################
    tk.Label(content_frame, text="Clínica/Escuela:").grid(row=row+0, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=school_var).grid(row=row+0, column=1, columnspan=2, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Fecha:").grid(row=row+0, column=4, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=date_var).grid(row=row+0, column=5, columnspan=2, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Nombre:").grid(row=row+1, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=name_var).grid(row=row+1, column=1, columnspan=2, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Fecha de nacimiento:").grid(row=row+1, column=4, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=birthday_var).grid(row=row+1, column=5, columnspan=3, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Número de teléfono:").grid(row=row+2, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=phone_number_var).grid(row=row+2, column=1, columnspan=2, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Razón por cita:").grid(row=row+3, column=0, padx=10, pady=5, sticky="nw")
    reason_for_visit_textbox = tk.Text(content_frame, height=5, width=60)
    reason_for_visit_textbox.grid(row=row+3, column=1, columnspan=6, padx=10, pady=5, sticky="w")

    ######################## vision distance ###############################
    tk.Label(content_frame, text="Visión distancia(dVA):").grid(row=row+4, column=0, padx=10, pady=5, sticky="w")
    vd_option1 = tk.Radiobutton(content_frame, text="s/c", variable=distance_vision_var, value="s/c")
    vd_option1.grid(row=row+4, column=1, padx=10, pady=5, sticky="w")
    vd_option2 = tk.Radiobutton(content_frame, text="c/c", variable=distance_vision_var, value="c/c")
    vd_option2.grid(row=row+4, column=2, padx=10, pady=5, sticky="w")

    vision_distance_validate_command = content_frame.register(utils.validate_vision_distance_input) 

    tk.Label(content_frame, text="OD:").grid(row=row+5, column=0, padx=10, pady=5, sticky="w")
    vd_entry_od = tk.Entry(content_frame, textvariable=distance_vision_OD_var, width=10, validate="key", validatecommand=(vision_distance_validate_command, "%P"))
    vd_entry_od.grid(row=row+5, column=1, padx=10, pady=5, sticky="w")
    vd_entry_od.insert(0, "20/") 

    tk.Label(content_frame, text="OI:").grid(row=row+6, column=0, padx=10, pady=5, sticky="w")
    vd_entry_oi = tk.Entry(content_frame, textvariable=distance_vision_OI_var, width=10, validate="key", validatecommand=(vision_distance_validate_command, "%P"))
    vd_entry_oi.grid(row=row+6, column=1, padx=10, pady=5, sticky="w")
    vd_entry_oi.insert(0, "20/") 


    ######################## intraocular pressure ###############################
    tk.Label(content_frame, text="Presión intraocular:").grid(row=row+4, column=4, padx=10, pady=5, sticky="w")
    ip_option1 = tk.Radiobutton(content_frame, text="iCare", variable=intraocular_pressure_var, value="iCare")
    ip_option1.grid(row=row+4, column=5, padx=10, pady=5, sticky="w")
    ip_option2 = tk.Radiobutton(content_frame, text="GAT", variable=intraocular_pressure_var, value="GAT")
    ip_option2.grid(row=row+4, column=6, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OD:").grid(row=row+5, column=4, padx=10, pady=5, sticky="w")
    ip_entry_od = tk.Entry(content_frame, textvariable=intraocular_pressure_OD_var, width=10)
    ip_entry_od.grid(row=row+5, column=5, padx=10, pady=5, sticky="w")
    mmhg_label_od = tk.Label(content_frame, text="mmHg")
    mmhg_label_od.grid(row=row+5, column=6, padx=0, pady=5, sticky="w")

    tk.Label(content_frame, text="OI:").grid(row=row+6, column=4, padx=10, pady=5, sticky="w")
    ip_entry_oi = tk.Entry(content_frame, textvariable=intraocular_pressure_OI_var, width=10)
    ip_entry_oi.grid(row=row+6, column=5, padx=10, pady=5, sticky="w")
    mmhg_label_oi = tk.Label(content_frame, text="mmHg")
    mmhg_label_oi.grid(row=row+6, column=6, padx=0, pady=5, sticky="w")

    ip_entry_od.bind("<FocusOut>", lambda event, text_var=intraocular_pressure_OD_var, label=mmhg_label_od, label_text="mmHg", timestamp_var=intraocular_pressure_timestamp_OD_var: utils.update_text_input_timestamp(label, label_text, timestamp_var, text_var))
    ip_entry_oi.bind("<FocusOut>", lambda event, text_var=intraocular_pressure_OI_var, label=mmhg_label_oi, label_text="mmHg", timestamp_var=intraocular_pressure_timestamp_OI_var: utils.update_text_input_timestamp(label, label_text, timestamp_var, text_var))


    ######################## dilatation ###############################
    dilatation_checkbox = tk.Checkbutton(content_frame, text="Dilatación", variable=dilatation_var)
    dilatation_checkbox.config(command=lambda checkbox_text="Dilatación", timestamp_var=dilatation_timestamp_var, checkbox_var=dilatation_var, checkbox=dilatation_checkbox: utils.update_checkbox_timestamp(timestamp_var, checkbox, checkbox_var, checkbox_text))
    dilatation_checkbox.grid(row=row+7, column=0, padx=5, pady=5, sticky="w")


    ######################## refractive status ###############################
    rs_option1 = tk.Radiobutton(content_frame, text="Autorefraction", variable=refractive_status_method_var, value="Autorefraction")
    rs_option1.grid(row=row+8, column=0, padx=0, pady=5, sticky="w")
    rs_option2 = tk.Radiobutton(content_frame, text="Retinoscopia", variable=refractive_status_method_var, value="Retinoscopia")
    rs_option2.grid(row=row+8, column=1, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OD:").grid(row=row+9, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=refractive_status_OD_var, width=10).grid(row=row+9, column=1, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="dVA:").grid(row=row+10, column=0, padx=20, pady=5, sticky="w")
    vd_entry_od = tk.Entry(content_frame, textvariable=refractive_status_OD_dVA_var, width=10, validate="key", validatecommand=(vision_distance_validate_command, "%P"))
    vd_entry_od.grid(row=row+10, column=1, padx=10, pady=5, sticky="w")
    vd_entry_od.insert(0, "20/") 

    tk.Label(content_frame, text="OI:").grid(row=row+11, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=refractive_status_OI_var, width=10).grid(row=row+11, column=1, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="dVA:").grid(row=row+12, column=0, padx=20, pady=5, sticky="w")
    vd_entry_oi = tk.Entry(content_frame, textvariable=refractive_status_OI_dVA_var, width=10, validate="key", validatecommand=(vision_distance_validate_command, "%P"))
    vd_entry_oi.grid(row=row+12, column=1, padx=10, pady=5, sticky="w")
    vd_entry_oi.insert(0, "20/") 


    ######################## lens prescription ###############################
    tk.Label(content_frame, text="Rx de anteojos").grid(row=row+8, column=4, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OD:").grid(row=row+9, column=4, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=lens_prescription_OD_var, width=10).grid(row=row+9, column=5, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OI:").grid(row=row+10, column=4, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=lens_prescription_OI_var, width=10).grid(row=row+10, column=5, padx=10, pady=5, sticky="w")


    ######################## table ###############################
    tk.Label(content_frame, text="OD").grid(row=row+13, column=0, columnspan=3)
    tk.Label(content_frame, text="OI").grid(row=row+13, column=4, columnspan=3)
    
    eyelid_OD_comment = tk.Entry(content_frame, textvariable=eyelid_OD_comments_var, width=10, state="disabled")
    eyelid_OI_comment = tk.Entry(content_frame, textvariable=eyelid_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Párpados", eyelid_OD_var, eyelid_OI_var, eyelid_OD_comment, eyelid_OI_comment, row+14)

    conjunctiva_OD_comment = tk.Entry(content_frame, textvariable=conjunctiva_OD_comments_var, width=10, state="disabled")
    conjunctiva_OI_comment = tk.Entry(content_frame, textvariable=conjunctiva_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Conjunctiva", conjunctiva_OD_var, conjunctiva_OI_var, conjunctiva_OD_comment, conjunctiva_OI_comment, row+15)

    cornea_OD_comment = tk.Entry(content_frame, textvariable=cornea_OD_comments_var, width=10, state="disabled")
    cornea_OI_comment = tk.Entry(content_frame, textvariable=cornea_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Cornea", cornea_OD_var, cornea_OI_var, cornea_OD_comment, cornea_OI_comment, row+16)

    iris_OD_comment = tk.Entry(content_frame, textvariable=iris_OD_comments_var, width=10, state="disabled")
    iris_OI_comment = tk.Entry(content_frame, textvariable=iris_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Iris", iris_OD_var, iris_OI_var, iris_OD_comment, iris_OI_comment, row+17)

    pupil_OD_comment = tk.Entry(content_frame, textvariable=pupil_OD_comments_var, width=10, state="disabled")
    pupil_OI_comment = tk.Entry(content_frame, textvariable=pupil_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Pupil", pupil_OD_var, pupil_OI_var, pupil_OD_comment, pupil_OI_comment, row+18)

    lente_OD_comment = tk.Entry(content_frame, textvariable=lente_OD_comments_var, width=10, state="disabled")
    lente_OI_comment = tk.Entry(content_frame, textvariable=lente_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Lente", lente_OD_var, lente_OI_var, lente_OD_comment, lente_OI_comment, row+19)

    tk.Label(content_frame, text="OD").grid(row=row+20, column=0, columnspan=3)
    tk.Label(content_frame, text="OI").grid(row=row+20, column=4, columnspan=3)

    tk.Entry(content_frame, textvariable=cd_OD_var, width=10).grid(row=row+21, column=0, columnspan=3, padx=10, pady=5, sticky="we")
    tk.Label(content_frame, text="C/D").grid(row=row+21, column=3, padx=5, pady=2)
    tk.Entry(content_frame, textvariable=cd_OI_var, width=10).grid(row=row+21, column=4, columnspan=3, padx=10, pady=5, sticky="we")

    nerve_OD_comment = tk.Entry(content_frame, textvariable=nerve_OD_comments_var, width=10, state="disabled")
    nerve_OI_comment = tk.Entry(content_frame, textvariable=nerve_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Nervio", nerve_OD_var, nerve_OI_var, nerve_OD_comment, nerve_OI_comment, row+22)

    macula_OD_comment = tk.Entry(content_frame, textvariable=macula_OD_comments_var, width=10, state="disabled")
    macula_OI_comment = tk.Entry(content_frame, textvariable=macula_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Macula", macula_OD_var, macula_OI_var, macula_OD_comment, macula_OI_comment, row+23)

    vasculature_OD_comment = tk.Entry(content_frame, textvariable=vasculature_OD_comments_var, width=10, state="disabled")
    vasculature_OI_comment = tk.Entry(content_frame, textvariable=vasculature_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Vasculatura", vasculature_OD_var, vasculature_OI_var, vasculature_OD_comment, vasculature_OI_comment, row+24)

    periphery_OD_comment = tk.Entry(content_frame, textvariable=periphery_OD_comments_var, width=10, state="disabled")
    periphery_OI_comment = tk.Entry(content_frame, textvariable=periphery_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Periferia", periphery_OD_var, periphery_OI_var, periphery_OD_comment, periphery_OI_comment, row+25)

    vitreous_OD_comment = tk.Entry(content_frame, textvariable=vitreous_OD_comments_var, width=10, state="disabled")
    vitreous_OI_comment = tk.Entry(content_frame, textvariable=vitreous_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Vitreous", vitreous_OD_var, vitreous_OI_var, vitreous_OD_comment, vitreous_OI_comment, row+26)


    ######################## diagnosis ###############################
    tk.Label(content_frame, text="Dx:").grid(row=row+27, column=0, sticky="w")
    tk.Checkbutton(content_frame, text="Miopía", variable=dx_myopia_var).grid(row=row+27, column=1, sticky="w")
    tk.Checkbutton(content_frame, text="Hipermetropía", variable=dx_hyperopia_var).grid(row=row+27, column=2, sticky="w")
    tk.Checkbutton(content_frame, text="Astigmatismo", variable=dx_astigmatism_var).grid(row=row+27, column=3, sticky="w")
    dx_comentarios_textbox = tk.Text(content_frame, height=3, width=60)
    dx_comentarios_textbox.grid(row=row+28, column=1, columnspan=6, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="Tx:").grid(row=row+29, column=0, sticky="nw")
    tx_textbox = tk.Text(content_frame, height=3, width=60)
    tx_textbox.grid(row=row+29, column=1, columnspan=6, padx=10, pady=5, sticky="w")

    ######################## doctor information ###############################
    tk.Label(content_frame, text="Firme de Doctor:").grid(row=row+30, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=doctor_firm_var, width=10).grid(row=row+30, column=1, columnspan=2, padx=10, pady=5, sticky="we")
    tk.Label(content_frame, text="Nombre de Doctor:").grid(row=row+31, column=0, padx=10, pady=5, sticky="w")
    tk.Entry(content_frame, textvariable=doctor_name_var, width=10).grid(row=row+31, column=1, columnspan=2, padx=10, pady=5, sticky="we")
    creation_date_var.set(datetime.now().strftime("%Y:%m:%d %H:%M:%S"))
    date = utils.format_date(creation_date_var.get())
    tk.Label(content_frame, text=f"Fecha: {date}").grid(row=row+31, column=6, padx=10, pady=5, sticky="w")


    # Save button
    save_button = tk.Button(content_frame, text="Save", command=lambda: save_to_excel(workbook, file_path, mmhg_label_od, mmhg_label_oi, 
                                                                        dilatation_checkbox, reason_for_visit_textbox, 
                                                                        dx_comentarios_textbox, tx_textbox,
                                                                        eyelid_OD_comment, eyelid_OI_comment,
                                                                        conjunctiva_OD_comment, conjunctiva_OD_comment, 
                                                                        cornea_OD_comment, cornea_OI_comment, 
                                                                        iris_OD_comment, iris_OI_comment, 
                                                                        pupil_OD_comment, pupil_OI_comment, 
                                                                        lente_OD_comment, lente_OI_comment, 
                                                                        nerve_OD_comment, nerve_OI_comment, 
                                                                        macula_OD_comment, macula_OI_comment, 
                                                                        vasculature_OD_comment, vasculature_OI_comment, 
                                                                        periphery_OD_comment, periphery_OI_comment, 
                                                                        vitreous_OD_comment, vitreous_OI_comment))
    save_button.grid(row=row+32, column=0, columnspan=7, pady=10)

    utils.apply_font_to_all_widgets(content_frame, 12)

def data_fill_form(window, workbook, content_frame, file_path, row, data, row_index):
    ######################## student information ###############################
    tk.Label(content_frame, text="Clínica/Escuela:").grid(row=row+0, column=0, padx=10, pady=5, sticky="w")
    school_var.set(data["Clínica/Escuela"])
    tk.Entry(content_frame, textvariable=school_var, state="readonly").grid(row=row+0, column=1, columnspan=2, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Fecha:").grid(row=row+0, column=4, padx=10, pady=5, sticky="w")
    date_var.set(data["Fecha"])
    tk.Entry(content_frame, textvariable=date_var, state="readonly").grid(row=row+0, column=5, columnspan=2, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Nombre:").grid(row=row+1, column=0, padx=10, pady=5, sticky="w")
    name_var.set(data["Nombre"])
    tk.Entry(content_frame, textvariable=name_var, state="readonly").grid(row=row+1, column=1, columnspan=2, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Fecha de nacimiento:").grid(row=row+1, column=4, padx=10, pady=5, sticky="w")
    birthday_var.set(data["Fecha_de_nacimiento"])
    tk.Entry(content_frame, textvariable=birthday_var, state="readonly").grid(row=row+1, column=5, columnspan=3, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Número de teléfono:").grid(row=row+2, column=0, padx=10, pady=5, sticky="w")
    phone_number_var.set(data["Número_de_teléfono"])
    tk.Entry(content_frame, textvariable=phone_number_var, state="readonly").grid(row=row+2, column=1, columnspan=2, padx=10, pady=5, sticky="we")
    
    reason_for_visit_text = data.get("Razón_por_cita", "")
    num_lines = len(reason_for_visit_text) // 60 + (1 if len(reason_for_visit_text) % 60 > 0 else 0)
    height = max(num_lines, 5)

    tk.Label(content_frame, text="Razón por cita:").grid(row=row+3, column=0, padx=10, pady=5, sticky="nw")
    reason_for_visit_textbox = tk.Text(content_frame, height=height, width=60)
    reason_for_visit_textbox.insert(tk.END, reason_for_visit_text)
    reason_for_visit_textbox.config(state="disabled") 
    reason_for_visit_textbox.grid(row=row+3, column=1, columnspan=6, padx=10, pady=5, sticky="w")

    ######################## vision distance ###############################
    tk.Label(content_frame, text="Visión distancia(dVA):").grid(row=row+4, column=0, padx=10, pady=5, sticky="w")
    distance_vision_var.set(data["Visión_distancia"])
    vd_option1 = tk.Radiobutton(content_frame, text="s/c", variable=distance_vision_var, value="s/c")
    vd_option1.grid(row=row+4, column=1, padx=10, pady=5, sticky="w")
    vd_option2 = tk.Radiobutton(content_frame, text="c/c", variable=distance_vision_var, value="c/c")
    vd_option2.grid(row=row+4, column=2, padx=10, pady=5, sticky="w")
    vd_option1.config(state="disabled")
    vd_option2.config(state="disabled")

    tk.Label(content_frame, text="OD:").grid(row=row+5, column=0, padx=10, pady=5, sticky="w")
    distance_vision_OD_var.set(data["Visión_distancia_OD"])
    vd_entry_od = tk.Entry(content_frame, textvariable=distance_vision_OD_var, width=10, state="readonly")
    vd_entry_od.grid(row=row+5, column=1, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OI:").grid(row=row+6, column=0, padx=10, pady=5, sticky="w")
    distance_vision_OI_var.set(data["Visión_distancia_OI"])
    vd_entry_oi = tk.Entry(content_frame, textvariable=distance_vision_OI_var, width=10, state="readonly")
    vd_entry_oi.grid(row=row+6, column=1, padx=10, pady=5, sticky="w")


    ######################## intraocular pressure ###############################
    tk.Label(content_frame, text="Presión intraocular:").grid(row=row+4, column=4, padx=10, pady=5, sticky="w")
    intraocular_pressure_var.set(data["Presión_intraocular"])
    ip_option1 = tk.Radiobutton(content_frame, text="iCare", variable=intraocular_pressure_var, value="iCare")
    ip_option1.grid(row=row+4, column=5, padx=10, pady=5, sticky="w")
    ip_option2 = tk.Radiobutton(content_frame, text="GAT", variable=intraocular_pressure_var, value="GAT")
    ip_option2.grid(row=row+4, column=6, padx=10, pady=5, sticky="w")
    ip_option1.config(state="disabled")
    ip_option2.config(state="disabled")

    tk.Label(content_frame, text="OD:").grid(row=row+5, column=4, padx=10, pady=5, sticky="w")
    intraocular_pressure_OD_var.set(data["Presión_intraocular_OD"])
    ip_entry_od = tk.Entry(content_frame, textvariable=intraocular_pressure_OD_var, width=10, state="readonly")
    ip_entry_od.grid(row=row+5, column=5, padx=10, pady=5, sticky="w")
    ip_OD_timestamp = data["Presión_intraocular_OD_tiempo"]
    mmhg_label_od = tk.Label(content_frame, text=f"mmHg @ {ip_OD_timestamp}")
    mmhg_label_od.grid(row=row+5, column=6, padx=0, pady=5, sticky="w")

    tk.Label(content_frame, text="OI:").grid(row=row+6, column=4, padx=10, pady=5, sticky="w")
    intraocular_pressure_OI_var.set(data["Presión_intraocular_OI"])
    ip_entry_oi = tk.Entry(content_frame, textvariable=intraocular_pressure_OI_var, width=10, state="readonly")
    ip_entry_oi.grid(row=row+6, column=5, padx=10, pady=5, sticky="w")
    ip_OI_timestamp = data["Presión_intraocular_OI_tiempo"]
    mmhg_label_oi = tk.Label(content_frame, text=f"mmHg @ {ip_OI_timestamp}")
    mmhg_label_oi.grid(row=row+6, column=6, padx=0, pady=5, sticky="w")

    ######################## dilatation ###############################
    dilatation_var.set(data["Dilatación"])
    dilatation_timestamp = data["Dilatación_tiempo"]
    dilatation_checkbox = tk.Checkbutton(content_frame, text=f"Dilatación @ {dilatation_timestamp}", variable=dilatation_var, state="disabled")
    dilatation_checkbox.grid(row=row+7, column=0, padx=5, pady=5, sticky="w")

    ######################## refractive status ###############################
    refractive_status_method_var.set(data["Autorefraction/Retinoscopia"])
    rs_option1 = tk.Radiobutton(content_frame, text="Autorefraction", variable=refractive_status_method_var, value="Autorefraction")
    rs_option1.grid(row=row+8, column=0, padx=0, pady=5, sticky="w")
    rs_option2 = tk.Radiobutton(content_frame, text="Retinoscopia", variable=refractive_status_method_var, value="Retinoscopia")
    rs_option2.grid(row=row+8, column=1, padx=10, pady=5, sticky="w")
    rs_option1.config(state="disabled")
    rs_option2.config(state="disabled")

    tk.Label(content_frame, text="OD:").grid(row=row+9, column=0, padx=10, pady=5, sticky="w")
    refractive_status_OD_var.set(data["Autorefraction/Retinoscopia_OD"])
    tk.Entry(content_frame, textvariable=refractive_status_OD_var, width=10, state="readonly").grid(row=row+9, column=1, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="dVA:").grid(row=row+10, column=0, padx=20, pady=5, sticky="w")
    refractive_status_OD_dVA_var.set(data["Autorefraction/Retinoscopia_OD_dVA"])
    vd_entry_od = tk.Entry(content_frame, textvariable=refractive_status_OD_dVA_var, width=10, state="readonly")
    vd_entry_od.grid(row=row+10, column=1, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OI:").grid(row=row+11, column=0, padx=10, pady=5, sticky="w")
    refractive_status_OI_var.set(data["Autorefraction/Retinoscopia_OI"])
    tk.Entry(content_frame, textvariable=refractive_status_OI_var, width=10, state="readonly").grid(row=row+11, column=1, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="dVA:").grid(row=row+12, column=0, padx=20, pady=5, sticky="w")
    refractive_status_OI_dVA_var.set(data["Autorefraction/Retinoscopia_OI_dVA"])
    vd_entry_oi = tk.Entry(content_frame, textvariable=refractive_status_OI_dVA_var, width=10, state="readonly")
    vd_entry_oi.grid(row=row+12, column=1, padx=10, pady=5, sticky="w")

    ######################## lens prescription ###############################
    tk.Label(content_frame, text="Rx de anteojos").grid(row=row+8, column=4, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OD:").grid(row=row+9, column=4, padx=10, pady=5, sticky="w")
    lens_prescription_OD_var.set(data["Rx_de_anteojos_OD"])
    tk.Entry(content_frame, textvariable=lens_prescription_OD_var, width=10, state="readonly").grid(row=row+9, column=5, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="OI:").grid(row=row+10, column=4, padx=10, pady=5, sticky="w")
    lens_prescription_OI_var.set(data["Rx_de_anteojos_OI"])
    tk.Entry(content_frame, textvariable=lens_prescription_OI_var, width=10, state="readonly").grid(row=row+10, column=5, padx=10, pady=5, sticky="w")


    ######################## table ###############################
    tk.Label(content_frame, text="OD").grid(row=row+13, column=0, columnspan=3)
    tk.Label(content_frame, text="OI").grid(row=row+13, column=4, columnspan=3)
    eyelid_OD_var.set(data["Párpados_OD"])
    eyelid_OD_comments_var.set(data["Párpados_OD_comentarios"] if data["Párpados_OD_comentarios"] is not None else "")
    eyelid_OI_var.set(data["Párpados_OI"])
    eyelid_OI_comments_var.set(data["Párpados_OI_comentarios"] if data["Párpados_OI_comentarios"] is not None else "")
    eyelid_OD_comment = tk.Entry(content_frame, textvariable=eyelid_OD_comments_var, width=10, state="disabled")
    eyelid_OI_comment = tk.Entry(content_frame, textvariable=eyelid_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Párpados", eyelid_OD_var, eyelid_OI_var, eyelid_OD_comment, eyelid_OI_comment, row+14, "disabled")
    
    conjunctiva_OD_var.set(data["Conjunctiva_OD"])
    conjunctiva_OD_comments_var.set(data["Conjunctiva_OD_comentarios"] if data["Conjunctiva_OD_comentarios"] is not None else "")
    conjunctiva_OI_var.set(data["Conjunctiva_OI"])
    conjunctiva_OI_comments_var.set(data["Conjunctiva_OI_comentarios"] if data["Conjunctiva_OI_comentarios"] is not None else "")
    conjunctiva_OD_comment = tk.Entry(content_frame, textvariable=conjunctiva_OD_comments_var, width=10, state="disabled")
    conjunctiva_OI_comment = tk.Entry(content_frame, textvariable=conjunctiva_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Conjunctiva", conjunctiva_OD_var, conjunctiva_OI_var, conjunctiva_OD_comment, conjunctiva_OI_comment, row+15, "disabled")

    cornea_OD_var.set(data["Cornea_OD"])
    cornea_OD_comments_var.set(data["Cornea_OD_comentarios"] if data["Cornea_OD_comentarios"] is not None else "")
    cornea_OI_var.set(data["Cornea_OI"])
    cornea_OI_comments_var.set(data["Cornea_OI_comentarios"] if data["Cornea_OI_comentarios"] is not None else "")
    cornea_OD_comment = tk.Entry(content_frame, textvariable=cornea_OD_comments_var, width=10, state="disabled")
    cornea_OI_comment = tk.Entry(content_frame, textvariable=cornea_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Cornea", cornea_OD_var, cornea_OI_var, cornea_OD_comment, cornea_OI_comment, row+16, "disabled")

    iris_OD_var.set(data["Iris_OD"])
    iris_OD_comments_var.set(data["Iris_OD_comentarios"] if data["Iris_OD_comentarios"] is not None else "")
    iris_OI_var.set(data["Iris_OI"])
    iris_OI_comments_var.set(data["Iris_OI_comentarios"] if data["Iris_OI_comentarios"] is not None else "")
    iris_OD_comment = tk.Entry(content_frame, textvariable=iris_OD_comments_var, width=10, state="disabled")
    iris_OI_comment = tk.Entry(content_frame, textvariable=iris_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Iris", iris_OD_var, iris_OI_var, iris_OD_comment, iris_OI_comment, row+17, "disabled")
    
    pupil_OD_var.set(data["Pupil_OD"])
    pupil_OD_comments_var.set(data["Pupil_OD_comentarios"] if data["Pupil_OD_comentarios"] is not None else "")
    pupil_OI_var.set(data["Pupil_OI"])
    pupil_OI_comments_var.set(data["Pupil_OI_comentarios"] if data["Pupil_OI_comentarios"] is not None else "")
    pupil_OD_comment = tk.Entry(content_frame, textvariable=pupil_OD_comments_var, width=10, state="disabled")
    pupil_OI_comment = tk.Entry(content_frame, textvariable=pupil_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Pupil", pupil_OD_var, pupil_OI_var, pupil_OD_comment, pupil_OI_comment, row+18, "disabled")

    lente_OD_var.set(data["Lente_OD"])
    lente_OD_comments_var.set(data["Lente_OD_comentarios"] if data["Lente_OD_comentarios"] is not None else "")
    lente_OI_var.set(data["Lente_OI"])
    lente_OI_comments_var.set(data["Lente_OI_comentarios"] if data["Lente_OI_comentarios"] is not None else "")
    lente_OD_comment = tk.Entry(content_frame, textvariable=lente_OD_comments_var, width=10, state="disabled")
    lente_OI_comment = tk.Entry(content_frame, textvariable=lente_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Lente", lente_OD_var, lente_OI_var, lente_OD_comment, lente_OI_comment, row+19, "disabled")
    
    tk.Label(content_frame, text="OD").grid(row=row+20, column=0, columnspan=3)
    tk.Label(content_frame, text="OI").grid(row=row+20, column=4, columnspan=3)

    cd_OD_var.set(data["C/D_OD"])
    cd_OI_var.set(data["C/D_OI"])
    tk.Entry(content_frame, textvariable=cd_OD_var, width=10, state="readonly").grid(row=row+21, column=0, columnspan=3, padx=10, pady=5, sticky="we")
    tk.Label(content_frame, text="C/D").grid(row=row+21, column=3, padx=5, pady=2)
    tk.Entry(content_frame, textvariable=cd_OI_var, width=10, state="readonly").grid(row=row+21, column=4, columnspan=3, padx=10, pady=5, sticky="we")

    nerve_OD_var.set(data["Nervio_OD"])
    nerve_OD_comments_var.set(data["Nervio_OD_comentarios"] if data["Nervio_OD_comentarios"] is not None else "")
    nerve_OI_var.set(data["Nervio_OI"])
    nerve_OI_comments_var.set(data["Nervio_OI_comentarios"] if data["Nervio_OI_comentarios"] is not None else "")
    nerve_OD_comment = tk.Entry(content_frame, textvariable=nerve_OD_comments_var, width=10, state="disabled")
    nerve_OI_comment = tk.Entry(content_frame, textvariable=nerve_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Nervio", nerve_OD_var, nerve_OI_var, nerve_OD_comment, nerve_OI_comment, row+22, "disabled")
    
    macula_OD_var.set(data["Macula_OD"])
    macula_OD_comments_var.set(data["Macula_OD_comentarios"] if data["Macula_OD_comentarios"] is not None else "")
    macula_OI_var.set(data["Macula_OI"])
    macula_OI_comments_var.set(data["Macula_OI_comentarios"] if data["Macula_OI_comentarios"] is not None else "")
    macula_OD_comment = tk.Entry(content_frame, textvariable=macula_OD_comments_var, width=10, state="disabled")
    macula_OI_comment = tk.Entry(content_frame, textvariable=macula_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Macula", macula_OD_var, macula_OI_var, macula_OD_comment, macula_OI_comment, row+23, "disabled")
    
    vasculature_OD_var.set(data["Vasculatura_OD"])
    vasculature_OD_comments_var.set(data["Vasculatura_OD_comentarios"] if data["Vasculatura_OD_comentarios"] is not None else "")
    vasculature_OI_var.set(data["Vasculatura_OI"])
    vasculature_OI_comments_var.set(data["Vasculatura_OI_comentarios"] if data["Vasculatura_OI_comentarios"] is not None else "")
    vasculature_OD_comment = tk.Entry(content_frame, textvariable=vasculature_OD_comments_var, width=10, state="disabled")
    vasculature_OI_comment = tk.Entry(content_frame, textvariable=vasculature_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Vasculatura", vasculature_OD_var, vasculature_OI_var, vasculature_OD_comment, vasculature_OI_comment, row+24, "disabled")
    
    periphery_OD_var.set(data["Periferia_OD"])
    periphery_OD_comments_var.set(data["Periferia_OD_comentarios"] if data["Periferia_OD_comentarios"] is not None else "")
    periphery_OI_var.set(data["Periferia_OI"])
    periphery_OI_comments_var.set(data["Periferia_OI_comentarios"] if data["Periferia_OI_comentarios"] is not None else "")
    periphery_OD_comment = tk.Entry(content_frame, textvariable=periphery_OD_comments_var, width=10, state="disabled")
    periphery_OI_comment = tk.Entry(content_frame, textvariable=periphery_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Periferia", periphery_OD_var, periphery_OI_var, periphery_OD_comment, periphery_OI_comment, row+25, "disabled")
    
    vitreous_OD_var.set(data["Vitreous_OD"])
    vitreous_OD_comments_var.set(data["Vitreous_OD_comentarios"] if data["Vitreous_OD_comentarios"] is not None else "")
    vitreous_OI_var.set(data["Vitreous_OI"])
    vitreous_OI_comments_var.set(data["Vitreous_OI_comentarios"] if data["Vitreous_OI_comentarios"] is not None else "")
    vitreous_OD_comment = tk.Entry(content_frame, textvariable=vitreous_OD_comments_var, width=10, state="disabled")
    vitreous_OI_comment = tk.Entry(content_frame, textvariable=vitreous_OI_comments_var, width=10, state="disabled")
    utils.create_row(content_frame, "Vitreous", vitreous_OD_var, vitreous_OI_var, vitreous_OD_comment, vitreous_OI_comment, row+26, "disabled")

    ######################## diagnosis ###############################
    tk.Label(content_frame, text="Dx:").grid(row=row+27, column=0, sticky="w")
    dx_myopia_var.set(data["Dx_Miopía"])
    dx_hyperopia_var.set(data["Dx_Hipermetropía"])
    dx_astigmatism_var.set(data["Dx_Astigmatismo"])
    tk.Checkbutton(content_frame, text="Miopía", variable=dx_myopia_var, state="disabled").grid(row=row+27, column=1, sticky="w")
    tk.Checkbutton(content_frame, text="Hipermetropía", variable=dx_hyperopia_var, state="disabled").grid(row=row+27, column=2, sticky="w")
    tk.Checkbutton(content_frame, text="Astigmatismo", variable=dx_astigmatism_var, state="disabled").grid(row=row+27, column=3, sticky="w")
    
    dx_comments_text = data["Dx_comentarios"]
    dx_comments_textbox = tk.Text(content_frame, height=3, width=60)
    dx_comments_textbox.insert(tk.END, dx_comments_text)
    dx_comments_textbox.config(state="disabled") 
    dx_comments_textbox.grid(row=row+28, column=1, columnspan=6, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="Tx:").grid(row=row+29, column=0, sticky="nw")
    tx_text = data["Tx"]
    tx_textbox = tk.Text(content_frame, height=3, width=60)
    tx_textbox.insert(tk.END, tx_text)
    tx_textbox.config(state="disabled")
    tx_textbox.grid(row=row+29, column=1, columnspan=6, padx=10, pady=5, sticky="w")

    ######################## doctor information ###############################
    tk.Label(content_frame, text="Firme de Doctor:").grid(row=row+30, column=0, padx=10, pady=5, sticky="w")
    doctor_firm_var.set(data["Firme_de_Doctor"])
    tk.Entry(content_frame, textvariable=doctor_firm_var, width=10, state="readonly").grid(row=row+30, column=1, columnspan=2, padx=10, pady=5, sticky="we")
    tk.Label(content_frame, text="Nombre de Doctor:").grid(row=row+31, column=0, padx=10, pady=5, sticky="w")
    doctor_name_var.set(data["Nombre_de_Doctor"])
    tk.Entry(content_frame, textvariable=doctor_name_var, width=10, state="readonly").grid(row=row+31, column=1, columnspan=2, padx=10, pady=5, sticky="we")
    creation_date = utils.format_date(data["Fecha_de_creación"])
    tk.Label(content_frame, text=f"Fecha de creación: {creation_date}").grid(row=row+30, column=5, columnspan=2, padx=10, pady=5, sticky="w")
    last_edit_date = utils.format_date(data["Fecha_de_la_última_edición"])
    tk.Label(content_frame, text=f"Fecha de la última edición: {last_edit_date}").grid(row=row+31, column=5, columnspan=2, padx=10, pady=5, sticky="w")

    tk.Label(content_frame, text="Addendum:").grid(row=row+32, column=0, columnspan=7, padx=10, pady=5, sticky="we")

    previous_addendums = data["Addendum"] if data["Addendum"] is not None else []
    if previous_addendums:
        addendum_list = previous_addendums.split("\n")
        for i, addendum in enumerate(addendum_list):
            tk.Label(content_frame, text=addendum, wraplength=800, justify="left").grid(
                row=row+33+i, column=0, columnspan=7, padx=10, pady=2, sticky="w"
            )

    row = row+32+len(previous_addendums) 

    addendum_textbox = tk.Text(content_frame, height=6, width=60)
    addendum_textbox.grid(row=row+1, column=0, columnspan=7, padx=10, pady=5, sticky="we")

    tk.Label(content_frame, text="Doctor Name:").grid(row=row+2, column=0, sticky="w", padx=5)
    addendum_doctor_name_entry = tk.Entry(content_frame)
    addendum_doctor_name_entry.grid(row=row+2, column=1, columnspan=2, padx=5, pady=5, sticky="w")

    addendum_date = datetime.now().strftime("%Y/%m/%d")
    tk.Label(content_frame, text=f"Date: {addendum_date}").grid(row=row+2, column=3, sticky="w", padx=5)

    tk.Button(content_frame, text="Close", command=lambda: on_window_close()).grid(row=row+3, column = 0, columnspan=3, pady=10, sticky="e")
    save_button = tk.Button(content_frame, text="Save", command=lambda: save_to_excel(workbook, file_path, mmhg_label_od, mmhg_label_oi, dilatation_checkbox, 
                                                                        reason_for_visit_textbox, dx_comments_textbox, tx_textbox, 
                                                                        eyelid_OD_comment, eyelid_OI_comment,
                                                                        conjunctiva_OD_comment, conjunctiva_OD_comment, 
                                                                        cornea_OD_comment, cornea_OI_comment, 
                                                                        iris_OD_comment, iris_OI_comment, 
                                                                        pupil_OD_comment, pupil_OI_comment, 
                                                                        lente_OD_comment, lente_OI_comment, 
                                                                        nerve_OD_comment, nerve_OI_comment, 
                                                                        macula_OD_comment, macula_OI_comment, 
                                                                        vasculature_OD_comment, vasculature_OI_comment, 
                                                                        periphery_OD_comment, periphery_OI_comment, 
                                                                        vitreous_OD_comment, vitreous_OI_comment, 
                                                                        addendum_textbox, addendum_doctor_name_entry, addendum_date, previous_addendums, row_index, window))
    save_button.grid(row=row+3, column=4, columnspan=3, pady=10, sticky="w")

    utils.apply_font_to_all_widgets(content_frame, 12)

    def on_window_close():
        clear_form(mmhg_label_od, mmhg_label_oi, dilatation_checkbox, reason_for_visit_textbox, dx_comments_textbox, tx_textbox, 
                    eyelid_OD_comment, eyelid_OI_comment, 
                    conjunctiva_OD_comment, conjunctiva_OD_comment, 
                    cornea_OD_comment, cornea_OI_comment, 
                    iris_OD_comment, iris_OI_comment, 
                    pupil_OD_comment, pupil_OI_comment, 
                    lente_OD_comment, lente_OI_comment, 
                    nerve_OD_comment, nerve_OI_comment, 
                    macula_OD_comment, macula_OI_comment, 
                    vasculature_OD_comment, vasculature_OI_comment, 
                    periphery_OD_comment, periphery_OI_comment, 
                    vitreous_OD_comment, vitreous_OI_comment)
        window.destroy() 

    window.protocol("WM_DELETE_WINDOW", on_window_close)